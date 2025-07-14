# generador_excel.py
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import math
import io # Importante para manejar el archivo en memoria

# --- Constantes y Diccionarios (Copiados directamente de tu script) ---
CM_TO_POINTS = 28.3461
ROW_1_HEIGHT_CM = 1.85
ROW_4_HEIGHT_CM = 2.8
DEFAULT_COLUMN_WIDTH = 8
HEADER_DATA_START_ROW = 4
NEW_DATA_START_ROW = 5
DEFAULT_GROSOR_ORILLA = 1.50
DEFAULT_GROSOR_CENTRO = 2.00
UMBRAL_ORILLA_CENTRO = 0.25
GRADOS_CIRCULO_COMPLETO = 360.0
INDICE_REFRACCION_MIN = 1.0
INDICE_REFRACCION_MAX = 2.0
RADIO_MINIMO_VALIDO = 0.01

ENTRADAS_FILA1_LABELS = {
    'B': "Num. Radios", 'C': "Índice Material (n)", 'D': "Esfera (D)",
    'E': "Cilindro (D)", 'F': "Eje Cilindro (°)", 'G': "Gr. Borde Min (mm)",
    'H': "Gr. Centro Min (mm)", 'I': "Gr. Referencia", 'J': "Decent. Horiz. (mm)",
    'K': "Decent. Vert. (mm)", 'L': "Magnitud Prisma (Δ)", 'M': "Base Prisma (°)",
    'N': "RX mas Positiva", 'O': "Sagita mas Positiva", 'P': "Grosor Prisma Máx.",
    'Q': "Radio Descent. Máx.", 'R': "PRISMA Grosor base-apice",
    'S': "Umbral Esf. (D)", 'T': "Gr. Borde Final Máx.", 'U': "Grosor Centro Final"
}
FORMATOS_NUMERO_FILA2 = {
    'B': '0', 'C': '0.000', 'D': '0.00', 'E': '0.00', 'F': '0', 'G': '0.00', 'H': '0.00',
    'I': '0.00', 'J': '0.00', 'K': '0.00', 'L': '0.00', 'M': '0', 'Q': '0.00',
    'R': '0.0000', 'N': '0.0000', 'O': '0.0000', 'P': '0.0000', 'S': '0.00',
    'T': '0.00', 'U': '0.00'
}
CELDAS_INPUT_DIRECTO_FILA2 = ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'J', 'K', 'L', 'M', 'S']
HEADERS_DATA_ROWS_TEXT = [
    'Radios', 'ANGULO Escrito', 'Ang. RADIANES', 'COSENO', 'SENO', 'X horiz', 'Y vert',
    'X-Desc.Horiz', 'Y-Desc.Vert', 'Radio-Desc.', 'Ang.RAD Desc.', 'RX en eje',
    'Sagita mm (Receta)', 'Gr.Orilla (Receta)', 'Grosor Prisma (Real)',
    'Grosor Orilla (N+O)', 'Grosor Orilla (N+O+I3)', 'Radio 2da img',
    'X Radio2 horiz', 'Y Radio2 vert', 'Aux: X Rotado Eje Prisma', 'Aux: Y Rotado Eje Prisma'
]
FORMULAS_TEMPLATE_DATA_ROWS = {
    'C': '=RADIANS(B{row})', 'D': '=COS(C{row})', 'E': '=SIN(C{row})',
    'F': '=$A{row}*COS(RADIANS($B{row}))', 'G': '=$A{row}*SIN(RADIANS($B{row}))',
    'H': '=IF(OR($A$1="R", $A$1="L"), IF($A$1="R", F{row}-$J$2, F{row}+$J$2), 0)',
    'I': '=G{row}-$K$2', 'J': '=SQRT(H{row}^2+I{row}^2)',
    'K': '=IF(ATAN2(I{row}, H{row}) < 0, ATAN2(I{row}, H{row}) + 2*PI(), ATAN2(I{row}, H{row}))',
    'L': '=$D$2+($E$2*POWER(SIN(RADIANS(ABS(DEGREES(K{row})-$F$2))),2))',
    'M': '=IFERROR(1000*(ABS(($C$2-1)*(1/IF(L{row}=0,0.00001,L{row})))-(SQRT(MAX(0,POWER(ABS(($C$2-1)*(1/IF(L{row}=0,0.00001,L{row})) ),2)-POWER((J{row}/1000),2)))))*SIGN(L{row}),0)',
    'N': '=IF($O$2<0,0,$O$2)-M{row}',
    'U': '=H{row}*COS(RADIANS(-$M$2)) - I{row}*SIN(RADIANS(-$M$2))',
    'V': '=H{row}*SIN(RADIANS(-$M$2)) + I{row}*COS(RADIANS(-$M$2))',
    'O': '=IF($R$2=0, 0, $R$2 * (U{row} - $AA$1) / ($AB$1 - $AA$1))',
    'P': '=N{row}+O{row}', 'Q': '=P{row}+$I$3', 'R': '=$A{row}+Q{row}',
    'S': '=$R{row}*COS(RADIANS($B{row}))', 'T': '=$R{row}*SIN(RADIANS($B{row}))',
}
NUMBER_FORMATS_DATA_ROWS = {
    'A': '0.00', 'B': '0.0', 'C': '0.0000', 'D': '0.0000', 'E': '0.0000',
    'F': '0.0000', 'G': '0.0000', 'H': '0.0000', 'I': '0.0000', 'J': '0.0000',
    'K': '0.0000', 'L': '0.0000', 'M': '0.0000', 'N': '0.0000', 'O': '0.0000',
    'P': '0.0000', 'Q': '0.00', 'R': '0.0000', 'S': '0.0000', 'T': '0.0000',
    'U': '0.0000', 'V': '0.0000'
}

# --- Todas las funciones de ayuda _... se mantienen igual ---
def _validar_datos_lente(datos_lente_dict):
    # ... (código idéntico a tu script) ...
    campos_requeridos = ["lado_ojo", "esfera_d", "cilindro_d", "eje_cilindro_grados", "prisma_magnitud_dp", "prisma_eje_base_grados", "indice_refraccion", "radios_borde_centesimas_mm_str", "decentracion_co_horizontal_mm", "decentracion_co_vertical_mm"]
    errores = []
    for campo in campos_requeridos:
        if campo not in datos_lente_dict:
            errores.append(f"Campo faltante: {campo}")
    if "indice_refraccion" in datos_lente_dict:
        indice = datos_lente_dict["indice_refraccion"]
        try:
            indice_float = float(indice)
            if not (INDICE_REFRACCION_MIN < indice_float < INDICE_REFRACCION_MAX):
                errores.append(f"Índice de refracción inválido: {indice}")
        except (ValueError, TypeError):
            errores.append(f"Índice de refracción no numérico: {indice}")
    if "lado_ojo" in datos_lente_dict:
        lado = datos_lente_dict["lado_ojo"]
        if lado not in ["R", "L", "r", "l"]:
            errores.append(f"Lado del ojo inválido: {lado}")
    if "eje_cilindro_grados" in datos_lente_dict:
        try:
            eje = float(datos_lente_dict["eje_cilindro_grados"])
            if not (0 <= eje <= GRADOS_CIRCULO_COMPLETO):
                errores.append(f"Eje cilindro fuera de rango: {eje}")
        except (ValueError, TypeError):
            errores.append(f"Eje cilindro no numérico: {datos_lente_dict['eje_cilindro_grados']}")
    if "prisma_eje_base_grados" in datos_lente_dict:
        try:
            eje_prisma = float(datos_lente_dict["prisma_eje_base_grados"])
            if not (0 <= eje_prisma <= GRADOS_CIRCULO_COMPLETO):
                errores.append(f"Eje base prisma fuera de rango: {eje_prisma}")
        except (ValueError, TypeError):
            errores.append(f"Eje base prisma no numérico: {datos_lente_dict['prisma_eje_base_grados']}")
    if errores:
        raise ValueError("Errores en datos de entrada:\n" + "\n".join(errores))
    return True

def _procesar_radios_seguros(radios_str_lista):
    # ... (código idéntico a tu script) ...
    radios_valores_mm = []
    for i, r in enumerate(radios_str_lista):
        if r.strip():
            try:
                valor = float(r) / 100.0
                if valor > RADIO_MINIMO_VALIDO:
                    radios_valores_mm.append(valor)
            except ValueError:
                pass
    return radios_valores_mm

def _inicializar_libro_y_hoja(titulo_hoja="Calculo"):
    # ... (código idéntico a tu script) ...
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = titulo_hoja
    return wb, ws

def _definir_estilos():
    # ... (código idéntico a tu script) ...
    return { "font_bold_black_size9": Font(bold=True, color="000000", size=9), "font_bold_black_size10": Font(bold=True, color="000000", size=10), "font_bold_black_size12": Font(bold=True, color="000000", size=12), "font_black_size8_center": Font(color="000000", size=8), "font_black_size10_center": Font(color="000000", size=10, bold=True), "font_black_size12_center": Font(color="000000", size=12, bold=True), "input_fill": PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"), "header_fill_new_row4": PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid"), "center_align_wrap": Alignment(horizontal='center', vertical='center', wrap_text=True), "right_align": Alignment(horizontal='right', vertical='center'), "center_align": Alignment(horizontal='center', vertical='center'), "border_thin_all": Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin')), "border_thick_all": Border(left=Side(style='thick'), right=Side(style='thick'), top=Side(style='thick'), bottom=Side(style='thick')) }

def _apply_block_border(ws, cell_range, border_style):
    # ... (código idéntico a tu script) ...
    rows = ws[cell_range]
    start_col, end_col = rows[0][0].column, rows[0][-1].column
    start_row, end_row = rows[0][0].row, rows[-1][0].row
    for row in rows:
        for cell in row:
            cell.border = Border()
    for cell in ws[start_row][start_col - 1:end_col]:
        cell.border += Border(top=border_style.top)
    for cell in ws[end_row][start_col - 1:end_col]:
        cell.border += Border(bottom=border_style.bottom)
    for row in ws.iter_rows(min_row=start_row, max_row=end_row, min_col=start_col, max_col=start_col):
        for cell in row:
            cell.border += Border(left=border_style.left)
    for row in ws.iter_rows(min_row=start_row, max_row=end_row, min_col=end_col, max_col=end_col):
        for cell in row:
            cell.border += Border(right=border_style.right)

def _configurar_encabezado_principal(ws, datos_lente_dict, estilos, num_rows_to_generate):
    # ... (código idéntico a tu script) ...
    ws['A1'] = datos_lente_dict.get("lado_ojo", "R").upper()
    ws['A1'].fill = estilos["input_fill"]
    ws['A1'].alignment = estilos["center_align"]
    ws['A1'].font = Font(bold=True, color="000000", size=40)
    for col, text in ENTRADAS_FILA1_LABELS.items():
        ws[f'{col}1'].value = text
        ws[f'{col}1'].font = estilos["font_bold_black_size10"]
        ws[f'{col}1'].alignment = estilos["center_align_wrap"]
    ws['B2'] = num_rows_to_generate
    ws['C2'] = datos_lente_dict["indice_refraccion"]
    ws['D2'] = datos_lente_dict["esfera_d"]
    ws['E2'] = datos_lente_dict["cilindro_d"]
    ws['F2'] = datos_lente_dict["eje_cilindro_grados"]
    ws['J2'] = datos_lente_dict["decentracion_co_horizontal_mm"]
    ws['K2'] = datos_lente_dict["decentracion_co_vertical_mm"]
    ws['L2'] = datos_lente_dict["prisma_magnitud_dp"]
    ws['M2'] = datos_lente_dict["prisma_eje_base_grados"]
    ws['S2'] = UMBRAL_ORILLA_CENTRO
    grosor_orilla_input = datos_lente_dict.get("grosor_orilla_mm", 0)
    grosor_centro_input = datos_lente_dict.get("grosor_centro_mm", 0)
    ws['G2'] = grosor_orilla_input if grosor_orilla_input > 0 else DEFAULT_GROSOR_ORILLA
    ws['H2'] = grosor_centro_input if grosor_centro_input > 0 else DEFAULT_GROSOR_CENTRO
    ws['I2'] = '=IF(D2>S2, G2, H2)'
    ws['H3'] = "Factor Grosor (I3)"
    ws['I3'] = '=Z10'
    if num_rows_to_generate > 0:
        last_data_row = NEW_DATA_START_ROW + num_rows_to_generate - 1
        ws['N2'] = f'=IFERROR(MAX(L{NEW_DATA_START_ROW}:L{last_data_row}),"")'
        ws['O2'] = f'=IFERROR(MAX(M{NEW_DATA_START_ROW}:M{last_data_row}),"")'
        ws['P2'] = f'=IFERROR(MAX(O{NEW_DATA_START_ROW}:O{last_data_row}),"")'
        ws['Q2'] = f'=IFERROR(MAX(J{NEW_DATA_START_ROW}:J{last_data_row}),"")'
        ws['R2'] = f'=IF(AND(ISNUMBER(L2), ISNUMBER(C2), C2-1<>0), (L2 * ($AB$1 - $AA$1)) / 100, "")'
        ws['T2'] = f'=IFERROR(MAX(Q{NEW_DATA_START_ROW}:Q{last_data_row}),"")'
        ws['U2'] = '=Z11'
        ws['AA1'] = f'=IFERROR(MIN(U{NEW_DATA_START_ROW}:U{last_data_row}), 0)'
        ws['AA1'].number_format = '0.0000'
        ws['AA1'].font = Font(size=8, color="808080")
        ws['AB1'] = f'=IFERROR(MAX(U{NEW_DATA_START_ROW}:U{last_data_row}), 0)'
        ws['AB1'].number_format = '0.0000'
        ws['AB1'].font = Font(size=8, color="808080")
    for col in ENTRADAS_FILA1_LABELS.keys():
        ws[f'{col}2'].alignment = estilos["right_align"]
        if col in FORMATOS_NUMERO_FILA2:
            ws[f'{col}2'].number_format = FORMATOS_NUMERO_FILA2.get(col)
        if col in CELDAS_INPUT_DIRECTO_FILA2:
            ws[f'{col}2'].fill = estilos["input_fill"]
    ws['I3'].number_format = '0.00'
    ws['I3'].alignment = estilos["right_align"]
    ws['H3'].alignment = Alignment(horizontal='right', vertical='center', wrap_text=True)

def _crear_cuadro_analisis_grosor(ws, estilos, num_rows_to_generate):
    # ... (código idéntico a tu script) ...
    if num_rows_to_generate == 0: return
    ws.merge_cells('Y3:Z3')
    title_cell = ws['Y3']
    title_cell.value = "Análisis de Grosor"
    title_cell.font = estilos["font_bold_black_size12"]
    title_cell.alignment = estilos["center_align"]
    title_cell.fill = estilos["header_fill_new_row4"]
    last_data_row = NEW_DATA_START_ROW + num_rows_to_generate - 1
    analysis_data = { 4: ("Sagita Mínima (mm)", f'=IFERROR(MIN(M{NEW_DATA_START_ROW}:M{last_data_row}),"")'), 5: ("Radio Desc. Mín. (mm)", f'=IFERROR(MIN(J{NEW_DATA_START_ROW}:J{last_data_row}),"")'), 6: ("Gr. Borde Mín. (sin I3)", f'=IFERROR(MIN(P{NEW_DATA_START_ROW}:P{last_data_row}),"")'), 7: ("Gr. Borde Final Mín. (mm)", f'=IFERROR(MIN(Q{NEW_DATA_START_ROW}:Q{last_data_row}),"")'), 8: ("Grosor Centro (sin I3)", '=IF(D2>S2, ABS(O2+(R2/2)), R2/2)'), 9: ("Proyección Ápice (mm)", f'=IFERROR(MIN(U{NEW_DATA_START_ROW}:U{last_data_row}),"")'), 10: ("Factor Grosor (I3)", '=MAX(I2-Z8, I2-Z6)'), 11: ("Grosor Centro Final (mm)", '=Z8+Z10') }
    for row_idx, (label, formula) in analysis_data.items():
        ws[f'Y{row_idx}'].value = label
        ws[f'Y{row_idx}'].font = Font(size=9)
        ws[f'Y{row_idx}'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        ws[f'Z{row_idx}'].value = formula
        ws[f'Z{row_idx}'].font = Font(size=9, bold=True)
        ws[f'Z{row_idx}'].number_format = '0.0000'
        ws[f'Z{row_idx}'].alignment = estilos["right_align"]
    ws['Z5'].number_format = '0.00'
    ws['Z6'].number_format = '0.00'
    ws['Z7'].number_format = '0.00'
    ws['Z8'].number_format = '0.00'
    ws['Z9'].number_format = '0.00'
    ws['Z10'].number_format = '0.00'
    ws['Z11'].number_format = '0.00'
    _apply_block_border(ws, 'Y3:Z11', estilos["border_thick_all"])
    ws.column_dimensions['Y'].width = 25
    ws.column_dimensions['Z'].width = 15

def _configurar_encabezados_datos(ws, estilos):
    # ... (código idéntico a tu script) ...
    for i, header_text in enumerate(HEADERS_DATA_ROWS_TEXT, 1):
        col_letter = get_column_letter(i)
        cell = ws[f'{col_letter}{HEADER_DATA_START_ROW}']
        cell.value = header_text
        cell.font = estilos["font_black_size10_center"]
        cell.fill = estilos["header_fill_new_row4"]
        cell.alignment = estilos["center_align_wrap"]
        cell.border = estilos["border_thin_all"]

def _configurar_dimensiones(ws):
    # ... (código idéntico a tu script) ...
    for i in range(1, len(HEADERS_DATA_ROWS_TEXT) + 1):
        ws.column_dimensions[get_column_letter(i)].width = DEFAULT_COLUMN_WIDTH
    ws.row_dimensions[1].height = ROW_1_HEIGHT_CM * CM_TO_POINTS
    ws.row_dimensions[HEADER_DATA_START_ROW].height = ROW_4_HEIGHT_CM * CM_TO_POINTS

def _poblar_filas_de_datos(ws, radios_valores_mm, num_rows_to_generate, estilos):
    # ... (código idéntico a tu script) ...
    if num_rows_to_generate == 0: return
    all_cols_to_process = set(FORMULAS_TEMPLATE_DATA_ROWS.keys()) | set(NUMBER_FORMATS_DATA_ROWS.keys())
    for i in range(num_rows_to_generate):
        crow = NEW_DATA_START_ROW + i
        ws[f'A{crow}'] = radios_valores_mm[i]
        ws[f'A{crow}'].number_format = NUMBER_FORMATS_DATA_ROWS['A']
        ws[f'A{crow}'].alignment = estilos["right_align"]
        if i == 0:
            ws[f'B{crow}'] = 0.0
        else:
            prev_b_val = ws[f'B{crow-1}'].value or 0
            ws[f'B{crow}'] = prev_b_val + (GRADOS_CIRCULO_COMPLETO / num_rows_to_generate)
        ws[f'B{crow}'].number_format = NUMBER_FORMATS_DATA_ROWS['B']
        ws[f'B{crow}'].alignment = estilos["right_align"]
        for col_letter in all_cols_to_process:
            if col_letter in ['A', 'B']: continue
            cell = ws[f'{col_letter}{crow}']
            if col_letter in FORMULAS_TEMPLATE_DATA_ROWS:
                cell.value = FORMULAS_TEMPLATE_DATA_ROWS[col_letter].format(row=crow)
            if col_letter in NUMBER_FORMATS_DATA_ROWS:
                cell.number_format = NUMBER_FORMATS_DATA_ROWS[col_letter]
            cell.alignment = estilos["right_align"]

def generar_excel_en_memoria(datos_lente_dict):
    """
    Función principal que toma los datos, crea el libro de Excel
    y lo devuelve como un objeto en memoria.
    """
    _validar_datos_lente(datos_lente_dict)

    lado_ojo = datos_lente_dict.get("lado_ojo", "R")
    nombre_etiqueta = datos_lente_dict["indice_refraccion"]
    wb, ws = _inicializar_libro_y_hoja(titulo_hoja=f"Calculo_{lado_ojo}_{nombre_etiqueta}")
    estilos = _definir_estilos()

    radios_str_lista = datos_lente_dict.get("radios_borde_centesimas_mm_str", "").split(';')
    radios_valores_mm = _procesar_radios_seguros(radios_str_lista)
    num_rows_to_generate = len(radios_valores_mm)

    _configurar_dimensiones(ws)
    _configurar_encabezado_principal(ws, datos_lente_dict, estilos, num_rows_to_generate)
    _configurar_encabezados_datos(ws, estilos)
    _poblar_filas_de_datos(ws, radios_valores_mm, num_rows_to_generate, estilos)
    _crear_cuadro_analisis_grosor(ws, estilos, num_rows_to_generate)

    _apply_block_border(ws, 'L1:M3', estilos["border_thick_all"])
    _apply_block_border(ws, 'J1:K3', estilos["border_thick_all"])
    _apply_block_border(ws, 'G1:I3', estilos["border_thick_all"])
    _apply_block_border(ws, 'D1:F3', estilos["border_thick_all"])
    _apply_block_border(ws, 'B1:C3', estilos["border_thick_all"])

    # --- CAMBIO CRÍTICO: Guardar en memoria en lugar de en un archivo ---
    buffer_virtual = io.BytesIO()
    wb.save(buffer_virtual)
    buffer_virtual.seek(0)
    return buffer_virtual.getvalue()